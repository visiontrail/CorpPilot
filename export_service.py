"""
Excel 导出服务模块
使用 openpyxl 在原 Excel 文件中追加分析结果 Sheet
保持原文件的样式和排版
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from typing import Dict, List
from io import BytesIO
import os


class ExcelExporter:
    """Excel 导出服务"""
    
    def __init__(self, file_path: str):
        """
        初始化导出服务
        
        Args:
            file_path: 原始 Excel 文件路径
        """
        self.file_path = file_path
        self.workbook = None
    
    def load_workbook(self):
        """加载原始 Excel 工作簿"""
        try:
            self.workbook = openpyxl.load_workbook(self.file_path)
        except Exception as e:
            raise ValueError(f"无法加载 Excel 文件: {str(e)}")
    
    def add_dashboard_sheet(self, dashboard_data: Dict):
        """
        添加 Dashboard_Data Sheet
        
        Args:
            dashboard_data: 分析结果字典
        """
        if self.workbook is None:
            self.load_workbook()
        
        # 删除已存在的同名 Sheet
        if "Dashboard_Data" in self.workbook.sheetnames:
            del self.workbook["Dashboard_Data"]
        
        # 创建新 Sheet
        ws = self.workbook.create_sheet("Dashboard_Data", 0)
        
        # 设置标题样式
        title_font = Font(name='微软雅黑', size=14, bold=True, color='FFFFFF')
        title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        title_alignment = Alignment(horizontal='center', vertical='center')
        
        # 设置内容样式
        content_font = Font(name='微软雅黑', size=11)
        content_alignment = Alignment(horizontal='left', vertical='center')
        
        # 边框样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        row_idx = 1
        
        # 1. KPI 汇总
        ws.merge_cells(f'A{row_idx}:F{row_idx}')
        cell = ws[f'A{row_idx}']
        cell.value = '📊 KPI 核心指标'
        cell.font = title_font
        cell.fill = title_fill
        cell.alignment = title_alignment
        row_idx += 1
        
        kpi = dashboard_data.get('kpi', {})
        kpi_data = [
            ['指标名称', '数值', '', '指标名称', '数值', ''],
            ['总差旅成本', f"¥{kpi.get('total_cost', 0):,.2f}", '', '总订单数', kpi.get('total_orders', 0), ''],
            ['异常记录数', kpi.get('anomaly_count', 0), '', '超标订单数', kpi.get('over_standard_count', 0), ''],
            ['紧急预订比例', f"{kpi.get('urgent_booking_ratio', 0)}%", '', '', '', '']
        ]
        
        for row_data in kpi_data:
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = content_font
                cell.alignment = content_alignment
                cell.border = thin_border
            row_idx += 1
        
        row_idx += 2
        
        # 2. 项目成本 Top 10
        ws.merge_cells(f'A{row_idx}:F{row_idx}')
        cell = ws[f'A{row_idx}']
        cell.value = '💰 项目成本 Top 10'
        cell.font = title_font
        cell.fill = title_fill
        cell.alignment = title_alignment
        row_idx += 1
        
        # 表头
        project_headers = ['项目代码', '总成本', '机票成本', '酒店成本', '火车票成本', '订单数量']
        for col_idx, header in enumerate(project_headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=header)
            cell.font = Font(name='微软雅黑', size=11, bold=True)
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            cell.alignment = title_alignment
            cell.border = thin_border
        row_idx += 1
        
        # 数据行
        top_projects = dashboard_data.get('top_projects', [])
        for project in top_projects:
            row_data = [
                project.get('项目代码', ''),
                project.get('总成本', 0),
                project.get('机票成本', 0),
                project.get('酒店成本', 0),
                project.get('火车票成本', 0),
                project.get('订单数量', 0)
            ]
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = content_font
                cell.alignment = content_alignment
                cell.border = thin_border
            row_idx += 1
        
        row_idx += 2
        
        # 3. 部门指标
        ws.merge_cells(f'A{row_idx}:F{row_idx}')
        cell = ws[f'A{row_idx}']
        cell.value = '🏢 部门指标汇总'
        cell.font = title_font
        cell.fill = title_fill
        cell.alignment = title_alignment
        row_idx += 1
        
        # 表头
        dept_headers = ['一级部门', '总成本', '总工时', '人员数量', '饱和度(%)', '']
        for col_idx, header in enumerate(dept_headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=header)
            cell.font = Font(name='微软雅黑', size=11, bold=True)
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            cell.alignment = title_alignment
            cell.border = thin_border
        row_idx += 1
        
        # 数据行
        dept_metrics = dashboard_data.get('department_metrics', [])
        for dept in dept_metrics:
            row_data = [
                dept.get('一级部门', ''),
                dept.get('总成本', 0),
                dept.get('总工时', 0),
                dept.get('人员数量', 0),
                dept.get('饱和度', 0),
                ''
            ]
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = content_font
                cell.alignment = content_alignment
                cell.border = thin_border
            row_idx += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
    
    def add_anomaly_sheet(self, anomalies: List[Dict]):
        """
        添加 Anomaly_Log Sheet
        
        Args:
            anomalies: 异常记录列表
        """
        if self.workbook is None:
            self.load_workbook()
        
        # 删除已存在的同名 Sheet
        if "Anomaly_Log" in self.workbook.sheetnames:
            del self.workbook["Anomaly_Log"]
        
        # 创建新 Sheet
        ws = self.workbook.create_sheet("Anomaly_Log")
        
        # 设置样式
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        content_font = Font(name='微软雅黑', size=10)
        content_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 表头
        headers = ['异常类型', '姓名', '日期', '考勤状态', '差旅类型', '差旅金额', '一级部门', '描述']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 数据行
        for row_idx, anomaly in enumerate(anomalies, start=2):
            row_data = [
                anomaly.get('Type', ''),
                anomaly.get('姓名', ''),
                anomaly.get('日期', ''),
                anomaly.get('考勤状态', ''),
                anomaly.get('差旅类型', ''),
                anomaly.get('差旅金额', 0),
                anomaly.get('一级部门', ''),
                anomaly.get('描述', '')
            ]
            
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = content_font
                cell.alignment = content_alignment
                cell.border = thin_border
                
                # 根据异常类型设置背景色
                if col_idx == 1:
                    if anomaly.get('Type') == 'Conflict':
                        cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    elif anomaly.get('Type') == 'NoExpense':
                        cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        
        # 调整列宽
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 40
        
        # 冻结首行
        ws.freeze_panes = 'A2'
    
    def save_to_bytes(self) -> BytesIO:
        """
        保存工作簿到内存字节流
        
        Returns:
            BytesIO 对象
        """
        if self.workbook is None:
            raise ValueError("工作簿未加载")
        
        output = BytesIO()
        self.workbook.save(output)
        output.seek(0)
        return output
    
    def save_to_file(self, output_path: str):
        """
        保存工作簿到文件
        
        Args:
            output_path: 输出文件路径
        """
        if self.workbook is None:
            raise ValueError("工作簿未加载")
        
        self.workbook.save(output_path)
    
    def export_with_analysis(
        self, 
        dashboard_data: Dict, 
        anomalies: List[Dict]
    ) -> BytesIO:
        """
        导出包含分析结果的 Excel 文件
        
        Args:
            dashboard_data: Dashboard 数据
            anomalies: 异常记录列表
            
        Returns:
            BytesIO 对象
        """
        self.load_workbook()
        self.add_dashboard_sheet(dashboard_data)
        self.add_anomaly_sheet(anomalies)
        return self.save_to_bytes()


